#!/usr/bin/env python3
"""Baut die aktualisierte Excel-Inventardatei (Stand: CPT-getriebenes System).

Quellen:
  scripts/data/vergleiche-master.json      – die 43 Vergleiche (Config/Status/Beschreibung)
  /tmp/staging-posts.json                  – aktuelle Beitrag→Vergleich-Verteilung
  scripts/output/live-embeds.json          – Live-Inventar (Herkunft)

Ausgabe: assets/beiträge/Vergleiche_Live_Inventar.xlsx (5 Blätter)
"""
import json, os, re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
OUT = os.path.join(ROOT, "assets", "beiträge", "Vergleiche_Live_Inventar.xlsx")

master = json.load(open(os.path.join(HERE, "data", "vergleiche-master.json"), encoding="utf-8"))
posts = json.load(open("/tmp/staging-posts.json", encoding="utf-8"))
live = json.load(open(os.path.join(HERE, "output", "live-embeds.json"), encoding="utf-8"))

# Vergleich-Slug -> Liste der Artikel, die ihn einbinden
cpt_to_posts = defaultdict(list)
for slug, d in posts.items():
    for v in d["vergleiche"]:
        cpt_to_posts[v].append(slug)

# Anbieter je Vergleich aus der Config ableiten
PROV = [("financeads.net", r"financeads"), ("Check24", r"check24|koop\.energie"), ("Covomo", r"covomo"),
        ("mr-money.de", r"mr-money"), ("partner-versicherung.de", r"partner-versicherung"),
        ("finanzen.de", r"fgrp|finanzen\.de"), ("bussgeldrechner.org", r"bussgeld"),
        ("Interhyp", r"interhyp"), ("InterRisk", r"interrisk|intervisio"),
        ("Die Haftpflichtkasse", r"haftpflichtkasse"), ("GDV/dieversicherer", r"dieversicherer")]
def provider(m):
    blob = (m.get("iframeUrl") or "") + json.dumps(m.get("scriptConfig") or {}) + (m.get("rawHtml") or "")
    for name, pat in PROV:
        if re.search(pat, blob, re.I): return name
    return m.get("provider") or "—"
def embed_target(m):
    if m.get("iframeUrl"): return m["iframeUrl"]
    if m.get("scriptConfig"): return "script:" + m["scriptConfig"].get("type", "")
    if m.get("rawHtml"): return "raw: " + m["rawHtml"][:60]
    return ""

HEADER_FILL = PatternFill("solid", fgColor="1F4E2E"); HEADER_FONT = Font(bold=True, color="FFFFFF")
OPEN_FILL = PatternFill("solid", fgColor="FCE8D5")
def style(ws, widths):
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]: c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions

wb = Workbook()

# ── Blatt 1: Vergleiche – aktueller Stand ────────────────────────────────────
ws = wb.active; ws.title = "Vergleiche – Stand"
ws.append(["Slug (CPT)", "Titel", "Anbieter", "Technik", "Status", "In Artikel(n)", "# Artikel",
           "Beschreibung (Textauszug)", "Widget-/Tool-URL"])
for m in sorted(master, key=lambda x: x["slug"]):
    arts = cpt_to_posts.get(m["slug"], [])
    ws.append([m["slug"], m.get("title", ""), provider(m), m.get("embedType", ""), m.get("status", ""),
               ", ".join(arts) if arts else "— OFFEN —", len(arts), m.get("desc", ""), embed_target(m)])
for row in ws.iter_rows(min_row=2):
    if row[5].value == "— OFFEN —":
        for c in row: c.fill = OPEN_FILL
style(ws, [42, 30, 22, 12, 9, 40, 8, 60, 58])

# ── Blatt 2: Beitrag → Vergleich (aktuelle Verteilung) ───────────────────────
ws2 = wb.create_sheet("Beitrag → Vergleich")
ws2.append(["Beitrag-Slug", "Titel", "Eingebundene Vergleiche", "Live-URL"])
for slug, d in sorted(posts.items()):
    if not d["vergleiche"]: continue
    ws2.append([slug, d["title"], ", ".join(d["vergleiche"]), f"https://staging.finanzleser.de/{slug}/"])
style(ws2, [38, 40, 50, 50])

# ── Blatt 3: Anbieter-Zusammenfassung ────────────────────────────────────────
ws3 = wb.create_sheet("Anbieter-Zusammenfassung")
ws3.append(["Anbieter", "# Vergleiche", "Technik(en)", "Vergleiche"])
byp = defaultdict(lambda: {"n": 0, "tech": set(), "slugs": []})
for m in master:
    p = provider(m); byp[p]["n"] += 1; byp[p]["tech"].add(m.get("embedType", "")); byp[p]["slugs"].append(m["slug"])
for p, d in sorted(byp.items(), key=lambda x: -x[1]["n"]):
    ws3.append([p, d["n"], ", ".join(sorted(d["tech"])), ", ".join(sorted(d["slugs"]))])
style(ws3, [24, 12, 24, 70])

# ── Blatt 4: Ohne Artikel / offen ────────────────────────────────────────────
ws4 = wb.create_sheet("Offen – ohne Artikel")
ws4.append(["Slug (CPT)", "Titel", "Anbieter", "Status", "Hinweis"])
for m in sorted(master, key=lambda x: x["slug"]):
    if cpt_to_posts.get(m["slug"]): continue
    ws4.append([m["slug"], m.get("title", ""), provider(m), m.get("status", ""),
                "Kein passender Beitrag vorhanden – manuell platzieren oder Beitrag anlegen"])
style(ws4, [42, 30, 22, 9, 60])

# ── Blatt 5: Live-Inventar (Herkunft, Referenz) ──────────────────────────────
ws5 = wb.create_sheet("Live-Inventar (Quelle)")
ws5.append(["Live-Beitrag", "Slug", "Anbieter", "Typ", "Technik", "Affiliate-ID", "Widget-URL", "Live-URL"])
for r in sorted(live, key=lambda x: (x.get("provider") or "zzz", x["slug"])):
    ws5.append([r["title"], r["slug"], r.get("provider") or "—", r["typ"], r["tech"],
                r["affiliate"], r["url"], r["link"]])
style(ws5, [38, 30, 22, 16, 14, 20, 55, 42])

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
placed = sum(1 for m in master if cpt_to_posts.get(m["slug"]))
print("Gespeichert:", OUT)
print(f"  Blatt 1: {len(master)} Vergleiche ({placed} in Artikeln, {len(master)-placed} offen)")
print(f"  Blatt 2: {sum(1 for d in posts.values() if d['vergleiche'])} Beiträge mit Vergleich")
print(f"  Blatt 3: {len(byp)} Anbieter | Blatt 4: {len(master)-placed} offen | Blatt 5: {len(live)} Live-Embeds")
